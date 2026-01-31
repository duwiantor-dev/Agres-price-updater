import io
import re
import zipfile
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# =========================
# Utils
# =========================
def s(x) -> str:
    return str(x).strip()


def sl(x) -> str:
    return str(x).strip().lower()


def normalize_int(val: Any) -> Optional[int]:
    """Ambil integer dari angka / teks '9.300' / '9,300' / '9300'."""
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    if isinstance(val, (int, float)):
        try:
            return int(round(float(val)))
        except Exception:
            return None
    txt = str(val)
    digits = re.sub(r"[^\d]", "", txt)
    return int(digits) if digits.isdigit() else None


def dedupe_columns(cols: List[Any]) -> List[Optional[str]]:
    out: List[Optional[str]] = []
    seen: Dict[str, int] = {}
    for c in cols:
        if c is None or (isinstance(c, float) and pd.isna(c)):
            out.append(None)
            continue
        name = s(c)
        if name not in seen:
            seen[name] = 0
            out.append(name)
        else:
            seen[name] += 1
            out.append(f"{name}_{seen[name]}")
    return out


def find_col_contains(df: pd.DataFrame, patterns: List[str]) -> Optional[str]:
    cols = list(df.columns)
    cols_l = [sl(c) for c in cols]
    for p in patterns:
        p_l = sl(p)
        for i, c_l in enumerate(cols_l):
            if p_l in c_l:
                return cols[i]
    return None


# =========================
# SKU parser
# =========================
def parse_platform_sku(full_sku: Any) -> Tuple[str, List[str]]:
    if full_sku is None or (isinstance(full_sku, float) and pd.isna(full_sku)):
        return "", []
    parts = s(full_sku).split("+")
    base = parts[0].strip()
    addons = [p.strip() for p in parts[1:] if p.strip()]
    return base, addons


# =========================
# Read Pricelist smart header
# =========================
def read_pricelist_smart(file, sheet_name=0) -> pd.DataFrame:
    raw = pd.read_excel(file, sheet_name=sheet_name, header=None)

    header_idx = None
    for i in range(min(25, len(raw))):
        row = [sl(x) for x in raw.iloc[i].tolist()]
        if "kodebarang" in row or "kode barang" in row:
            header_idx = i
            break

    if header_idx is None:
        st.error("❌ Header Pricelist tidak ditemukan (kolom KODEBARANG).")
        st.stop()

    df = raw.iloc[header_idx:].copy()
    df.columns = dedupe_columns(df.iloc[0].tolist())
    df = df.iloc[1:].reset_index(drop=True)
    df = df.loc[:, df.columns.notna()]
    return df


# =========================
# Scale auto-detect (AUTO x1000 atau tidak)
# =========================
def detect_price_scale_from_pricelist(df_pl: pd.DataFrame, price_col: str) -> int:
    nums = []
    for v in df_pl[price_col].head(200).tolist():
        iv = normalize_int(v)
        if iv is not None and iv > 0:
            nums.append(iv)

    if not nums:
        return 1000

    nums.sort()
    median = nums[len(nums) // 2]
    return 1000 if median < 100000 else 1


# =========================
# Build maps
# =========================
@dataclass
class Rules:
    price_scale: int
    discount_rp: int  # diskon langsung rupiah (final)


def build_base_maps(
    df_pl: pd.DataFrame,
    col_sku: str,
    col_price: str,
    col_stock_tot: str,
    price_scale: int
) -> Tuple[Dict[str, int], Dict[str, int]]:
    price_map: Dict[str, int] = {}
    stock_map: Dict[str, int] = {}

    for _, r in df_pl.iterrows():
        sku = s(r.get(col_sku, ""))
        if not sku:
            continue

        price_raw = normalize_int(r.get(col_price, None))
        if price_raw is not None:
            price_map[sku] = int(price_raw) * int(price_scale)

        stock_raw = normalize_int(r.get(col_stock_tot, None))
        if stock_raw is not None:
            stock_map[sku] = int(stock_raw)

    return price_map, stock_map


def build_addon_map(df_add: pd.DataFrame, price_scale: int) -> Dict[str, int]:
    """
    CASE-INSENSITIVE:
      - key disimpan uppercase
    """
    code_col = (
        find_col_contains(df_add, ["standarisasi kode sku di varian"]) or
        find_col_contains(df_add, ["addon_code"]) or
        find_col_contains(df_add, ["kode sku"]) or
        find_col_contains(df_add, ["kode"])
    )
    price_col = find_col_contains(df_add, ["harga"]) or find_col_contains(df_add, ["price"])

    if not code_col or not price_col:
        st.error("❌ Kolom kode/harga tidak ditemukan di Addon Mapping.")
        st.stop()

    m: Dict[str, int] = {}
    for _, r in df_add.iterrows():
        code = s(r.get(code_col, "")).upper()
        if not code:
            continue
        val = normalize_int(r.get(price_col, None))
        if val is None:
            continue
        m[code] = int(val) * int(price_scale)

    return m


def calc_addon_total(addons: List[str], addon_map: Dict[str, int]) -> Tuple[int, List[str]]:
    """
    CASE-INSENSITIVE:
      - token addon dibandingkan uppercase
    """
    total = 0
    missing = []
    for a in addons:
        key = s(a).upper()
        if key in addon_map:
            total += addon_map[key]
        else:
            missing.append(s(a))
    return total, missing


# =========================
# Mass Update: preserve format using openpyxl
# =========================
def find_header_row_and_cols(ws, scan_rows: int = 30) -> Tuple[int, Dict[str, int]]:
    for r in range(1, scan_rows + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        vals_l = [sl(v) for v in vals]

        hits = {}
        if "sku" in vals_l:
            hits["sku"] = vals_l.index("sku") + 1
        if "harga" in vals_l:
            hits["harga"] = vals_l.index("harga") + 1
        if "stok" in vals_l:
            hits["stok"] = vals_l.index("stok") + 1

        if "sku" in hits and "harga" in hits:
            return r, hits

    return -1, {}


def detect_marketplace_from_filename(filename: str) -> Optional[str]:
    name = sl(filename)
    if "shopee" in name:
        return "shopee"
    if "tiktok" in name:
        return "tiktok"
    return None


def process_one_mass_workbook(
    mass_bytes: bytes,
    base_price_map: Dict[str, int],
    base_stock_map: Dict[str, int],
    addon_map: Dict[str, int],
    rules: Rules
) -> Tuple[bytes, pd.DataFrame]:
    """
    RULE BARU (sesuai request kamu):
    - Jika base SKU tidak ketemu -> JANGAN UBAH APA PUN
    - Jika ada 1 addon saja yang tidak ketemu -> JANGAN UBAH APA PUN
    - Kalau aman (base ketemu & semua addon ketemu):
        HargaFinal = (base + addon) - diskon_rp
        Stok = TOT
    """
    wb = load_workbook(io.BytesIO(mass_bytes))
    ws = wb.worksheets[0]

    header_row, cols = find_header_row_and_cols(ws, scan_rows=30)
    if header_row == -1:
        raise ValueError("Header Mass Update tidak ketemu (butuh kolom sku & harga).")

    sku_col = cols["sku"]
    harga_col = cols["harga"]
    stok_col = cols.get("stok")

    issues = []
    empty_run = 0

    for r in range(header_row + 1, ws.max_row + 1):
        sku_val = ws.cell(row=r, column=sku_col).value

        if sku_val is None or s(sku_val) == "":
            empty_run += 1
            if empty_run >= 50:
                break
            continue
        empty_run = 0

        base, addons = parse_platform_sku(sku_val)
        base_price = base_price_map.get(base)

        # base tidak ketemu -> skip total
        if base_price is None:
            continue

        addon_total, missing = calc_addon_total(addons, addon_map)

        # addon ada yang missing -> jangan ubah apa pun
        if missing:
            issues.append({
                "row": r,
                "sku_full": s(sku_val),
                "base_sku": base,
                "reason": f"SKIP karena addon tidak ketemu: {','.join(missing)}"
            })
            continue

        final_price = int(base_price + addon_total - rules.discount_rp)
        if final_price < 0:
            final_price = 0

        ws.cell(row=r, column=harga_col).value = final_price

        if stok_col is not None:
            base_stock = base_stock_map.get(base)
            if base_stock is not None:
                ws.cell(row=r, column=stok_col).value = int(base_stock)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), pd.DataFrame(issues)


# =========================
# UI (rapih, minimal)
# =========================
st.set_page_config(page_title="Web App Update Harga", layout="wide")
st.title("Web App Update Harga")

c1, c2, c3 = st.columns(3)
with c1:
    mass_files = st.file_uploader("Upload Mass Update (bisa banyak)", type=["xlsx"], accept_multiple_files=True)
with c2:
    pl_file = st.file_uploader("Upload Pricelist", type=["xlsx"])
with c3:
    addon_file = st.file_uploader("Upload Addon Mapping", type=["xlsx"])

st.divider()

discount_input = st.number_input("Diskon (Rp) - mengurangi harga final", value=0, step=1000)

process_btn = st.button("Proses")


if process_btn:
    if not mass_files or pl_file is None or addon_file is None:
        st.warning("Upload Mass Update (minimal 1), Pricelist, dan Addon Mapping dulu ya.")
        st.stop()

    df_pl = read_pricelist_smart(pl_file, sheet_name=0)
    df_add = pd.read_excel(addon_file, sheet_name=0)

    col_sku = find_col_contains(df_pl, ["kodebarang"]) or find_col_contains(df_pl, ["kode barang"])
    col_tot = find_col_contains(df_pl, ["tot"])
    if not col_sku:
        st.error("❌ Pricelist: kolom KODEBARANG tidak ditemukan.")
        st.stop()
    if not col_tot:
        st.error("❌ Pricelist: kolom TOT tidak ditemukan.")
        st.stop()

    sample_price_col = find_col_contains(df_pl, ["m3"]) or find_col_contains(df_pl, ["m4"])
    if not sample_price_col:
        st.error("❌ Pricelist: kolom M3/M4 tidak ditemukan.")
        st.stop()

    auto_scale = detect_price_scale_from_pricelist(df_pl, sample_price_col)
    rules = Rules(price_scale=int(auto_scale), discount_rp=int(discount_input))

    addon_map = build_addon_map(df_add, price_scale=rules.price_scale)

    zip_buf = io.BytesIO()
    all_issues = []

    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for mf in mass_files:
            mp = detect_marketplace_from_filename(mf.name)
            if mp is None:
                mp = "tiktok"
                all_issues.append({
                    "file": mf.name,
                    "row": "",
                    "sku_full": "",
                    "base_sku": "",
                    "reason": "Nama file tidak mengandung 'tiktok'/'shopee' → default TikTok (M3)"
                })

            if mp == "tiktok":
                col_price = find_col_contains(df_pl, ["m3"])
                price_label = "M3"
            else:
                col_price = find_col_contains(df_pl, ["m4"])
                price_label = "M4"

            if not col_price:
                st.error(f"❌ Pricelist: kolom {price_label} tidak ditemukan.")
                st.stop()

            base_price_map, base_stock_map = build_base_maps(
                df_pl=df_pl,
                col_sku=col_sku,
                col_price=col_price,
                col_stock_tot=col_tot,
                price_scale=rules.price_scale
            )

            try:
                updated_bytes, report = process_one_mass_workbook(
                    mass_bytes=mf.getvalue(),
                    base_price_map=base_price_map,
                    base_stock_map=base_stock_map,
                    addon_map=addon_map,
                    rules=rules
                )
            except Exception as e:
                all_issues.append({
                    "file": mf.name,
                    "row": "",
                    "sku_full": "",
                    "base_sku": "",
                    "reason": f"Gagal proses file: {e}"
                })
                continue

            out_name = mf.name.replace(".xlsx", "_updated.xlsx")
            zf.writestr(out_name, updated_bytes)

            if not report.empty:
                rep2 = report.copy()
                rep2.insert(0, "file", mf.name)
                all_issues.extend(rep2.to_dict("records"))

    if all_issues:
        df_issues = pd.DataFrame(all_issues)
        issues_bytes = io.BytesIO()
        with pd.ExcelWriter(issues_bytes, engine="xlsxwriter") as w:
            df_issues.to_excel(w, index=False, sheet_name="issues")
        with zipfile.ZipFile(zip_buf, "a", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("issues_report.xlsx", issues_bytes.getvalue())

    st.success("✅ Selesai")
    st.download_button(
        label="⬇️ Download ZIP",
        data=zip_buf.getvalue(),
        file_name="mass_update_results.zip",
        mime="application/zip"
    )
